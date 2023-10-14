"""Desktop GUI attendance recording program
Every new day the application is launched, it writes the date in the next column for recording.
Ideally, the camera in place takes a picture of the student and passes the file path of the captured image and sends it to the program, after which it identifies the student whose image was detected and records their timestamp[arrival time of student].

Example:
John Doe    19:36:15
Jane Doe        -      
"""
# Python Imports
import datetime
from datetime import datetime
import openpyxl
import cv2 as cv
import numpy as np
import os
# Local Import
from facereg.tools import *



class Attendance():
    def __init__(self, students) -> None:
        # Attendance spreadsheet
        self.fp = os.path.abspath("atten.xlsx")
        # Spreadsheet workbook instance
        self.wb = openpyxl.load_workbook(self.fp)
        self.sheet = self.wb.active
        self.students = students

    def retrain(self) -> None:
        """Retrains the facial models and stores their coordinates for identification(testing)."""
        people = []
        dir = r'.\facereg\faces\train'
        for name in os.listdir(os.path.abspath(dir)):
            people.append(name)

        features = []
        labels = []
        # train loop for each person
        for person in people:
            path = os.path.join(dir, person)
            label = people.index(person)

            for img in os.listdir(path):
                img_path = os.path.join(path, img)

                img_array = cv.imread(img_path)
                gray = cv.cvtColor(img_array, cv.COLOR_BGR2GRAY)

                faces_rect = HAAR_CASCADE.detectMultiScale(gray, scaleFactor=1.1, minNeighbors=4)

                for (x,y,w,h) in faces_rect:
                    faces_roi = gray[y:y+h, x:x+w]
                    features.append(faces_roi)
                    labels.append(label)

        features = np.array(features, dtype='object')
        labels = np.array(labels)
        face_recog = cv.face.LBPHFaceRecognizer.create()

        face_recog.train(features, labels)

        face_recog.save(r'.\facereg\face_trained.yml')
        np.save(r'.\facereg\features.npy', features)
        np.save(r'.\facereg\labels.npy', labels)

        showdialog('Models retrained')

    def add_student(self, name) -> None:
        """Add a student to the attendance spreadsheet
        
        Keyword arguments:
        name -- Student's name
        Return: None
        """
        max_row = get_max_row(self.sheet)
        # Check if student exists
        for row in self.sheet.iter_rows(max_col=1, min_row=2, max_row=max_row):
            for cell in row:
                # Cell contains a value and that value is the student's name
                if cell.value is not None and name in cell.value:
                    showdialog('Student aleady exists.')
                    return
        self.sheet.cell(row=max_row+1, column=1).value = name
        self.wb.save(self.fp)
        showdialog('Student added.')

    def detect(self, image_path) -> int or bool:
        """Main image detection for each captured image
        
        Keyword arguments:
        image_path -- File path of captured image
        Return: Returns student label if face is detected, False otherwise.
        """
        # Loading student image
        img = cv.imread(image_path)

        # Switch image to gray scale and display it
        gray = cv.cvtColor(img, cv.COLOR_BGR2GRAY)
        # cv.imshow('gray', gray)

        # Detect face in image
        faces_rect = HAAR_CASCADE.detectMultiScale(gray, 1.1, 4)

        if faces_rect is None:
            return False
        else:
            # Draws rectangle on image face and displays label[name] on image
            for (x,y,w,h) in faces_rect:
                faces_roi = gray[y:y+h, x:x+h]

                label, confidence = FACE_RECOG.predict(faces_roi)
                # TODO: Use click echo since it's a CLI tool now

                # print(f'Label = {self.students[label]} with a confidence of {confidence}')

                cv.putText(img, str(self.students[label]), (20,20), cv.FONT_HERSHEY_COMPLEX, 1.0, (0, 255,0), thickness=2)
                cv.rectangle(img, (x,y), (x+w, y+h), (0,255,0), thickness=2)
            # Display image
            cv.imshow('detected face', img)
            cv.waitKey(0)
            return label, f'Label = {self.students[label]} with a confidence of {confidence}'

    def new_date(self) -> None:
        """Records the new day date in the next empty column
        
        Keyword arguments:
        argument -- description
        Return: return_description
        """        
        max_col = get_max_col(self.sheet)
        today = str(datetime.now().date())
        # print(f'Today\'s date is {today}')
        # If date has already been added, skip
        if self.sheet.cell(row=1, column=max_col).value == today:
            return
        else:
            self.sheet.cell(row=1, column=max_col+1).value = today
            showdialog('New date added.')
            self.wb.save(self.fp)

    def record_attendance(self, image_path) -> str:
        """Record attendance time in the next empty column
        
        Keyword arguments:
        self -- Instance of the class
        image_path -- File path of the captured image to be recognized
        max_col -- The last column with recorded times
        Return: `Student Not Found` or `[name] PRESENT: [time of arrival]`
        """
        detect_res = self.detect(image_path)
        if not detect_res:
            showdialog("Student not found 🚫")
            return
        else:
            student_label = detect_res[0]
            # Arrival record
            time = datetime.now().time()

            max_row = get_max_row(self.sheet)
            max_col = get_max_col(self.sheet)
            # Writing arrival record to next column
            for row in self.sheet.iter_rows(min_col=1, max_col=1, min_row=2, max_row=max_row):
                for cell in row:
                    # Cell contains a value and that value is the student's reg number
                    if cell.value is not None and self.students[student_label] in cell.value:
                        # Get row of student
                        stud_row = cell.row
                        # Write attendance timestamp
                        self.sheet.cell(row=stud_row, column=max_col).value = time
            self.wb.save(self.fp)
            showdialog('Arrival recorded ✔')
